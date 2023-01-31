import traceback
from urllib import response
import traceback
from datetime import datetime

import requests
from lxml.html import fromstring
import asyncio
from pyppeteer import launch
from pyppeteer_stealth import stealth
from lxml.html import fromstring
import json
from openpyxl.workbook import Workbook

urls = [line.rstrip('\n') for line in open('urls.txt') if len(line) > 1 and line[0] != '#']
from openpyxl.workbook import Workbook
import time

async def scrape():
    # iteration on urls.txt fille
    response = {}
    response["success"]=True

    for url in urls :
        #stop is true when collect data
        Stop=False
        # iteration when data from url is not collected
        while Stop==False:
            #initialize browser
            browser = await launch(headless=False, executablePath="C:/Program Files/Google/Chrome/Application/chrome.exe", handleSIGINT=False,
    handleSIGTERM=False,
    handleSIGHUP=False
                                   )
            #initialize page
            page = await browser.newPage()
            try:
                await stealth(page)  # <-- Here
                time.sleep(5)
                # go to url with default time out
                try:

                    await page.goto(url)
                except:
                    # if an error occured we wait 10 s until the page is loaded
                    time.sleep(10)
                    pass

                #get html content from url
                html = await page.content()
                #transforem html on wpath to acces data ith
                html_xpath = fromstring(html)

                # extract json where data to collect is stored
                info = html_xpath.xpath('//script[@type="application/ld+json"]/text()')[0]
                # load data as json to acces using key , value
                info = json.loads(info)
                # begin block get restaurant data

                rest_name = info["name"]


                rest_adress = info["address"]["streetAddress"]
                rest_city = info["address"]["addressLocality"]
                rest_state = info["address"]["addressRegion"]
                rest_stars = info["aggregateRating"]["ratingValue"]
                rest_review_count = info["aggregateRating"]["reviewCount"]
                response["rest_name"] = rest_name
                response["rest_adress"] = rest_adress
                response["rest_city"] = rest_city
                response["rest_state"] = rest_state
                response["rest_review_count"] = rest_review_count
                response["rest_stars"] = rest_stars
                print("Restaurant Name : ", rest_name)
                print("Restaurant Address Line 1 : ", rest_adress)
                print("Restaurant City  : ", rest_city)
                print("Restaurant State : ", rest_state)
                print("Restaurant Stars : ", rest_stars)
                print("Restaurant Review Count : ", rest_review_count)
                #end getting restaurant data
                menus = info["hasMenu"]["hasMenuSection"]
                wb = Workbook()
                #begin getting menus
                for menu in menus:
                    print("menus names : ", menu["name"])
                    #creating a sheet on csv file as name the menu name
                    ws1 = wb.create_sheet(menu["name"])

                    row = 1
                    column = 1

                    # preaparing columns names
                    header = ['Description', 'Price']

                    # writing columns names
                    for item in header:
                        # write operation perform
                        _ = ws1.cell(column=column, row=row, value="{0}".format(item))
                        column += 1
                    column = 1
                    for has_menu in menu["hasMenuItem"]:
                        #block preparing menues item (desccription and price)
                        row = row + 1
                        description = has_menu["description"]

                        price = has_menu["offers"]["price"]
                        if len(description) == 0 or len(description) == 1:
                            description = menu["name"]
                        content = [description, price]
                        column = 1
                        for item in content:
                            _ = ws1.cell(column=column, row=row, value="{0}".format(item))
                            column = column + 1
                #save csv file as name the restaurent name
                wb.save(filename='{0}.xlsx'.format(rest_name))
                # stop is true when data are collected end csv file is stored
                Stop = True
                await browser.close();
                return response
            except:
                # the exception block is used for retry when something go wrong
                try:
                    import traceback
                    print(traceback.format_exc())
                    print("here in exception")
                    time.sleep(10)
                    html = await page.content()
                    html_xpath = fromstring(html)
                    save_html_page(html);
                    info = html_xpath.xpath('//script[@type="application/ld+json"]/text()')[0]
                    if info:

                        info = json.loads(info)
                        rest_name=info["name"]
                        rest_adress=info["address"]["streetAddress"]
                        rest_city=info["address"]["addressLocality"]
                        rest_state=info["address"]["addressRegion"]
                        rest_stars=info["aggregateRating"]["ratingValue"]
                        rest_review_count=info["aggregateRating"]["reviewCount"]
                        response["rest_name"] = rest_name
                        response["rest_adress"] = rest_adress
                        response["rest_city"] = rest_city
                        response["rest_state"] = rest_state
                        response["rest_review_count"] = rest_review_count
                        response["rest_stars"] = rest_stars
                        print("Restaurant Name : ",rest_name)
                        print("Restaurant Address Line 1 : ",rest_adress)
                        print("Restaurant City  : ",rest_city)
                        print("Restaurant State : ",rest_state)
                        print("Restaurant Stars : ",rest_stars)
                        print("Restaurant Review Count : ",rest_review_count)
                        menus=info["hasMenu"]["hasMenuSection"]
                        wb = Workbook()
                        for menu in menus:
                            print("menus names : ", menu["name"])
                            ws1 = wb.create_sheet(menu["name"])

                            row = 1
                            column = 1

                            # preapartion des noms des colonnes por les informations produits
                            header = ['Description', 'Price']

                            # ecriture de noms des colonnes
                            for item in header:
                                # write operation perform
                                _ = ws1.cell(column=column, row=row, value="{0}".format(item))
                                column += 1
                            column = 1
                            for has_menu in menu["hasMenuItem"] :
                                row = row + 1
                                description=has_menu["description"]
                                price= has_menu["offers"]["price"]
                                if len(description)==0 or len (description)==1:
                                    description=menu["name"]
                                content = [description,price]
                                column=1
                                for item in content:
                                    _ = ws1.cell(column=column, row=row, value="{0}".format(item))
                                    column = column+1

                        wb.save(filename='{0}.xlsx'.format(rest_name))
                        Stop = True
                        await browser.close()
                        return response

                    else:

                        await browser.close()

                    return response
                except:
                    try:
                        print("hehehehehh")
                        await browser.close()


                    except:
                        pass
                return response
def save_html_page(page_source):
	helloFile = open('page.html', 'w')
	helloFile.write(str(page_source))
	print("saved")
	helloFile.close()






